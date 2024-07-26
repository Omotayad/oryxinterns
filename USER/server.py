import pandas as pd
import pickle
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import os
import openpyxl
import io
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sqlalchemy import create_engine
import time
import schedule
import logging
import threading
from functools import wraps
import jwt
import datetime
import requests
import bcrypt
from db_utils import get_user, check_password

# Set up logging
logging.basicConfig(level=logging.INFO)

# Database connection details
db_host = 'localhost'
db_name = 'NHP'
db_user = 'postgres'
db_password = 'Admin9114'


# Add this global variable to store active users
active_users = set()

# Create the database engine
engine = create_engine(f'postgresql://{db_user}:{db_password}@{db_host}/{db_name}')

# Load the pre-trained model and column transformer
dt_model = pickle.load(open("model/dt_model.pkl", "rb"))
column_transformer = pickle.load(open('model/column_transformer.pkl', 'rb'))

def retrain_model():
    global dt_model, column_transformer
    logging.info("Starting model retraining...")
    
    try:
        # Fetch new data from the database
        new_data = pd.read_sql_query("""
            SELECT
                itemname,
                doctype,
                linetotal,
                EXTRACT(YEAR FROM docdate) AS year,
                quantity,
                itemcost,
                grssprofit
            FROM product_sales
            WHERE EXTRACT(YEAR FROM docdate) NOT IN (2015, 2016, 2017, 2018, 2019, 2020, 2021)
        """, engine)

        # Preprocess the new data
        new_data = new_data[['itemname', 'doctype', 'linetotal', 'year', 'quantity', 'itemcost', 'grssprofit']]
        X_new = new_data[['itemname', 'doctype', 'linetotal', 'year', 'quantity', 'itemcost']]
        y_new = new_data['grssprofit']
        categorical_cols = ['itemname', 'doctype']
        numerical_cols = ['year', 'linetotal', 'quantity', 'itemcost']
        column_transformer = ColumnTransformer([('one_hot_encoder', OneHotEncoder(handle_unknown='ignore'), categorical_cols)], remainder='passthrough')
        X_new = column_transformer.fit_transform(X_new)

        logging.info(f"New data shape: {new_data.shape}")

        # Retrain the model with the new data
        dt_model.fit(X_new, y_new)
        
        logging.info(f"Model retrained. New score: {dt_model.score(X_new, y_new)}")

        # Save the updated model
        pickle.dump(dt_model, open("model/dt_model.pkl", "wb"))
        pickle.dump(column_transformer, open('model/column_transformer.pkl', 'wb'))
        
        logging.info("Model retraining completed and saved.")
    except Exception as e:
        logging.error(f"Error during model retraining: {str(e)}")

# Schedule the model retraining to run at 00:00 every day
schedule.every().day.at("00:00").do(retrain_model)

# Run model retraining immediately if it hasn't been run today
if not schedule.jobs:
    retrain_model()

# Function to run the scheduler in a separate thread
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(60)  # Check every minute

# Start the scheduler in a separate thread
scheduler_thread = threading.Thread(target=run_scheduler)
scheduler_thread.start()

app = Flask(__name__, static_url_path='', static_folder='.')
CORS(app)
app.config['SECRET_KEY'] = 'CNHP25'  # Change this to a secure random key

required_fields = ['itemname', 'doctype', 'linetotal', 'year', 'quantity', 'itemcost']

# Create a new Excel file for storing predictions
predictions_file = 'predictions.xlsx'
if not os.path.exists(predictions_file):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(['ItemName', 'DocType', 'LineTotal', 'Year', 'Quantity', 'ItemCost', 'Prediction'])
    workbook.save(predictions_file)

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'message': 'Token is missing!'}), 401
        try:
            data = jwt.decode(token.split()[1], app.config['SECRET_KEY'], algorithms=["HS256"])
        except:
            return jsonify({'message': 'Token is invalid!'}), 401
        return f(*args, **kwargs)
    return decorated

def generate_token(username):
    token = jwt.encode({
        'user': username,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24)
    }, app.config['SECRET_KEY'], algorithm="HS256")
    return token

def get_sales_trend(item_name):
    # Query the database to get sales trend data for all years
    query = """
    SELECT EXTRACT(YEAR FROM docdate) as year,
           SUM(linetotal) as total_sales
    FROM product_sales
    WHERE itemname = %s
    GROUP BY EXTRACT(YEAR FROM docdate)
    ORDER BY year
    """
    df = pd.read_sql_query(query, engine, params=(item_name,))
    
    return {
        'year': df['year'].tolist(),
        'total_sales': df['total_sales'].tolist()
    }

def notify_admin_server(username, role):
    try:
        admin_server_url = 'http://127.0.0.1:5002/admin/user-login'
        data = {
            'username': username,
            'role': role,
            'login_time': datetime.datetime.now().isoformat()
        }
        response = requests.post(admin_server_url, json=data)
        if response.status_code != 200:
            logging.error(f"Failed to notify admin server about user login: {response.text}")
    except Exception as e:
        logging.error(f"Error notifying admin server about user login: {str(e)}")

@app.route('/')
def serve_index():
    return send_from_directory('.', 'index.html')

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    user = get_user(username)
    if user and check_password(password, user['password']):
        token = jwt.encode({
            'user': username,
            'role': user['role'],
            'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm="HS256")
        active_users.add(username)

        # Notify admin server about the login
        notify_admin_server(username, user['role'])

        return jsonify({"message": "Login successful", "token": token, "role": user['role']}), 200
    
    return jsonify({"message": "Wrong Password or Username!"}), 401

# Add a new logout route
@app.route('/logout', methods=['POST'])
@token_required
def logout():
    token = request.headers.get('Authorization').split()[1]
    try:
        data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
        username = data['user']
        # Remove the user from active_users set
        active_users.discard(username)
        return jsonify({"message": "Logout successful"}), 200
    except:
        return jsonify({"message": "Invalid token"}), 401

# Add a new route to get active users count
@app.route('/active-users-count', methods=['GET'])
def get_active_users_count():
    return jsonify({"active_users_count": len(active_users)}), 200

@app.route('/predict', methods=['POST'])
@token_required
def predict():
    try:
        new_data = request.get_json()
        if new_data is None:
            return jsonify({'error': 'No data provided'}), 400
        
        missing_fields = [field for field in required_fields if field not in new_data]
        if missing_fields:
            return jsonify({'error': f'Missing fields: {missing_fields}'}), 400

        # Convert keys to lowercase
        new_data = {k.lower(): v for k, v in new_data.items()}
        
        # Create DataFrame with lowercase column names
        new_df = pd.DataFrame([new_data])
        
        # Ensure all required columns are present
        for col in ['itemname', 'doctype', 'linetotal', 'year', 'quantity', 'itemcost']:
            if col not in new_df.columns:
                new_df[col] = None

        new_data_transformed = column_transformer.transform(new_df)
        pred = dt_model.predict(new_data_transformed)

        # Store the prediction data in the Excel file
        workbook = openpyxl.load_workbook(predictions_file)
        worksheet = workbook.active
        worksheet.append([new_data[field.lower()] for field in required_fields] + [int(pred[0])])
        workbook.save(predictions_file)

        # Notify admin server about successful prediction
        try:
            requests.post('http://127.0.0.1:5002/admin/increment-predictions')
        except:
            logging.error("Failed to notify admin server about prediction")

        # Get the item name from the form data
        item_name = new_data['itemname']

        # Get updated chart data
        sales_trend = get_sales_trend(item_name)
        
        return jsonify({
            'prediction': int(pred[0]),
            'salesTrend': sales_trend
        })
        
    except Exception as e:
        logging.error(f"Error during prediction: {str(e)}")
        return jsonify({'error': 'An error occurred during prediction'}), 500


@app.route('/download_predictions', methods=['GET'])
@token_required
def download_predictions():
    try:
        workbook = openpyxl.load_workbook(predictions_file)
        data = workbook.active
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return send_file(
            stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='predictions.xlsx'
        )
    except Exception as e:
        logging.error(f"Error during prediction download: {str(e)}")
        return jsonify({'error': 'An error occurred while downloading predictions'}), 500

if __name__ == '__main__':
    # Run the Flask app in the main thread
    app.run(host='127.0.0.1', port=5001, debug=True, use_reloader=False)